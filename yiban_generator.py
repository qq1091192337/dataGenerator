import json
import os
import openpyxl
import datetime
import re
from os.path import *
import pandas as pd
import docx.document
from docx import Document
from docx.table import *

from docx.enum.table import *
from docx.oxml.shared import OxmlElement, qn
from openpyxl.cell import Cell
inputRoot='./input'
detectHeader={
    "序号":0,
    "专业":1,
    "班级":2,
    "姓名":3,
    "工号/学号":4,
    "人员类型":5,
    "性别":6,
    "年级":7,
    "手机号":8,
    "状态":9,
}
nameDict={}
dateDict={}
notifyDict={}
todayNotifyDict={}
data=[[],[],[],[],[],[],[],[],[]]
def preProcessCell(cellText :str,col :int):
    if col ==1 and '土木' in cellText:
        return "土木学院"
    if col ==6:
        matches_num = re.findall(r'^(\d{2})$',cellText)
        if matches_num:
            return f"{matches_num[0]}级"
        match =re.match(r'^((\d{2}){2})$',cellText)
        if match:
            return f"{match.group(1)}"
    # if '土木' in cellText:
    #     return "土木学院"
    return cellText.strip()

def initData():
    global nameDict
    if exists('./yiban/data.json'):
        with open('./yiban/data.json','r') as f:
            f.read(999999)
            nameDict = json.load(f)
def readInput():
    global detectHeader,nameDict
    inp= filter (lambda e:e.endswith(".xlsx"),os.listdir('./yiban/input'))
    for e in inp:
        workbook= openpyxl.load_workbook(f'./yiban/input/{e}')
        sheet1= workbook.worksheets[0]
        for row in sheet1.iter_rows():
            for cell in row:
                if cell.value in detectHeader:
                    detectHeader[cell.value]=cell.column
            if '姓名'==  row[detectHeader["姓名"]-1].value \
                    or None == row[detectHeader["姓名"]-1].value\
                    or '已删除'== row[detectHeader["姓名"]-1].value:
                continue
            if row[detectHeader["姓名"]-1].value in nameDict:
                nameDict[row[detectHeader["姓名"]-1].value].update( {
                    "专业":row[detectHeader["专业"]-1].value,
                    "班级":row[detectHeader["班级"]-1].value,
                    "姓名":row[detectHeader["姓名"]-1].value,
                    "人员类型":row[detectHeader["人员类型"]-1].value,
                    "工号/学号":row[detectHeader["工号/学号"]-1].value,
                    "性别":row[detectHeader["性别"]-1].value,
                    "年级":row[detectHeader["年级"]-1].value,
                    "手机号":row[detectHeader["手机号"]-1].value,
                    "状态":row[detectHeader["状态"]-1].value,
                    "次数":nameDict[row[detectHeader["姓名"]-1].value]['次数']+1
                })
                nameDict[row[detectHeader["姓名"]-1].value]['时间'].append(re.match(r".*(\d+月\d+日).+",e).group(1))
            else:
                nameDict[row[detectHeader["姓名"]-1].value] = {
                    "专业":row[detectHeader["专业"]-1].value,
                    "班级":row[detectHeader["班级"]-1].value,
                    "姓名":row[detectHeader["姓名"]-1].value,
                    "人员类型":row[detectHeader["人员类型"]-1].value,
                    "工号/学号":row[detectHeader["工号/学号"]-1].value,
                    "性别":row[detectHeader["性别"]-1].value,
                    "年级":row[detectHeader["年级"]-1].value,
                    "手机号":row[detectHeader["手机号"]-1].value,
                    "状态":row[detectHeader["状态"]-1].value,
                    "时间":[re.match(r".*(\d+月\d+日).+",e).group(1)],
                    "次数":1
                }


        pass
if __name__ == '__main__':
    inp= input("""请输入操作的类型：
    1：读取数据
    2：分析数据，生成的汇总表和通报批评名单
    3：根据名单生成通报批评文本
    4：根据分析数据生成工作简报表格
    """)

    if inp=='1':
        readInput()
        # with open('yiban/data.json','r') as f:
        #     try:
        #         nameDict.update(json.load(f))
        #     except json.decoder.JSONDecodeError:
        #         pass

        with open('yiban/data.json','w') as f:
            json.dump(nameDict,f)
        sumx=0
        for v in nameDict.values():
            sumx += v['次数']
        print(sumx)
    if inp=='2':

        with open('yiban/data.json','r') as f:
            nameDict.update(json.load(f))
        with open('yiban/notify.json','r') as f:
            last_notify = json.load(f)
        for k,v in nameDict.items():
            if v['次数']>=2:
                notifyDict[k]=v
                for e in v['时间']:
                    if e in dateDict:
                        dateDict[e].append(v)
                    else:
                        dateDict[e]=[v]
        todayNotifyList=[]
        for k,v in notifyDict.items():
            if k in last_notify:#说明不是第一次
                if last_notify[k]['次数']!=notifyDict[k]['次数']:
                    if notifyDict[k]['次数']%2==0:
                        todayNotifyList.append(notifyDict[k])
            else:
                todayNotifyList.append(notifyDict[k])
        if not exists(f'{datetime.datetime.now().strftime("%y-%m-%d")}通报.txt'):
            with open(f'{datetime.datetime.now().strftime("%y-%m-%d")}通报.txt','w') as f:
                txt = '\n'.join([x['姓名'] for x in todayNotifyList])
                f.writelines(txt)

        wb=openpyxl.Workbook()
        #ws=wb.create_sheet('Sheet1')
        ws=wb.active
        ws.merge_cells("A1:I2")
        ws.cell(1,1).value=datetime.datetime.today().strftime("3月16日-%m月%d日通报批评人员名单")
        ws.cell(3,1).value='专业班级'
        ws.cell(3, 2).value = '姓名'
        ws.cell(3, 3).value = '人员类型'
        ws.cell(3, 4).value = '学号'
        ws.cell(3, 5).value = '性别'
        ws.cell(3, 6).value = '年级'
        ws.cell(3, 7).value = '手机号码'
        ws.cell(3, 8).value = '累计次数'
        ws.cell(3,9).value='未打卡日期'
        pointer=4

        for e in notifyDict.values():
            ws.cell(pointer,1).value=e['班级']
            ws.cell(pointer,2).value=e['姓名']
            ws.cell(pointer,3).value=e['人员类型']
            ws.cell(pointer,4).value=e['工号/学号']
            ws.cell(pointer,5).value=e['性别']
            ws.cell(pointer,6).value=e['年级']
            ws.cell(pointer,7).value=e['手机号']
            ws.cell(pointer,8).value=e['次数']
            ws.cell(pointer,9).value=','.join(e['时间'])

            pointer+=1
        wb.save(f"./yiban/output/3月16日-{datetime.datetime.today().strftime('%m月%d日')}通报批评人员名单.xlsx")

        wb2 = openpyxl.Workbook()
        # ws=wb.create_sheet('Sheet1')
        ws1 = wb2.active
        ws1.merge_cells("A1:I2")
        ws1.cell(1, 1).value = datetime.datetime.today().strftime("3月16日-%m月%d日通报批评人员名单")
        ws1.cell(3, 1).value = '专业班级'
        ws1.cell(3, 2).value = '姓名'
        ws1.cell(3, 3).value = '人员类型'
        ws1.cell(3, 4).value = '学号'
        ws1.cell(3, 5).value = '性别'
        ws1.cell(3, 6).value = '年级'
        ws1.cell(3, 7).value = '手机号码'
        ws1.cell(3, 8).value = '累计次数'
        ws1.cell(3, 9).value = '未打卡日期'
        pointer = 4

        for e in nameDict.values():
            ws1.cell(pointer, 1).value = e['班级']
            ws1.cell(pointer, 2).value = e['姓名']
            ws1.cell(pointer, 3).value = e['人员类型']
            ws1.cell(pointer, 4).value = e['工号/学号']
            ws1.cell(pointer, 5).value = e['性别']
            ws1.cell(pointer, 6).value = e['年级']
            ws1.cell(pointer, 7).value = e['手机号']
            ws1.cell(pointer, 8).value = e['次数']
            ws1.cell(pointer, 9).value = ','.join(e['时间'])

            pointer += 1
        wb2.save(f"./yiban/output/土木学院未按时打卡汇总表（3月16日-{datetime.datetime.today().strftime('%m月-%d日')}）.xlsx")
        with open('yiban/notify.json', 'w') as f:
            json.dump(notifyDict,f)
        with open('yiban/date.json', 'w') as f:
            json.dump(dateDict,f)
    if inp=='3':
        with open('yiban/data.json', 'r') as f:
            nameDict.update(json.load(f))
        l = []
        while (True):
            x= input()

            if x=='ok' or x=='':
                break
            l.append(x)
        s=""
        d=[]

        nameDict_sort_by_times=sorted([nameDict[c] for c in l],key=lambda e :e['次数'],reverse=True)
        last_time=nameDict_sort_by_times[0]['次数']
        nameText=''
        df= pd.DataFrame(nameDict_sort_by_times)
        groups= df.groupby('次数',sort=False)
        for name,group in groups:
            for ele in group.iloc:
                nameText+=f'''{ele["姓名"]}、'''
            nameText= nameText.strip('、')
            nameText+=f"{name}次未按时打卡；"
        for e in nameDict_sort_by_times:
            s+=f"{e['姓名']}，{e['性别']}，"\
                               f"学号：{e['工号/学号']}，系{e['年级']}级{e['班级']}班学生；"
        #     if last_time==e['次数']:
        #         nameText+=f'{e["姓名"]}、'
        #     else:
        #         last_time=e['次数']
        #         nameText= nameText.strip('、')
        #         nameText+=f'{e["次数"]+1}次未按时打卡；'
        #         nameText += f'{e["姓名"]}、'
        # nameText= nameText.strip('、')
        # nameText+=f'{last_time}次未按时打卡；'
        nameText=nameText.strip('；')
        s+=f'''\n以上同学在学院 2022 年 3 月 16 日至 4 月 XX 日的每日健康打卡中，{nameText}。根据长沙理工大学《关于疫情期间进一步加强校园师生管理的紧急通知》【2021】第 11 号、《长沙理工大学学生违纪处理办法》 【2021】5 号相关规定，经学院学工办研究决定，给予{'、'.join(l)} {len(l)} 位同学通报批评。同时将以上通报结果纳入学年内综合测评及评奖推优考核。'''
        print(s)
    if inp=='4':
        with open('./yiban/date.json') as f:
            name_sort_by_date:dict =json.load(f)
        def genTimeRange():
            today= datetime.datetime.today()
            # if not today.weekday() ==6:
            #     raise Exception('当前不是星期日')
            lastDay=today+datetime.timedelta(days=-7)
            res=[]
            for i in range((today-lastDay).days+1):
                delta= lastDay + datetime.timedelta(days=i)
                res.append(f"{delta.month}月{delta.day}日")
            return res
        timeRange= genTimeRange()
        recordDict={}
        for e in timeRange:
            for v  in name_sort_by_date[e]:
                if v['姓名'] in recordDict:
                    recordDict[v['姓名']]['次数']+=1
                else:
                    recordDict[v['姓名']]=v
                    recordDict[v['姓名']]['次数']=1
        printList=[ v for k,v in recordDict.items() if v['次数']>=2]
        printList =sorted(printList,key=lambda x:x['次数'],reverse=True)
        wb2 = openpyxl.Workbook()
        # ws=wb.create_sheet('Sheet1')
        ws1 = wb2.active
        ws1.merge_cells("A1:I2")
        ws1.cell(1, 1).value = f'{timeRange[0]}-{timeRange[-1]}日通报批评名单\n备注："次数"为上述范围内的未打卡次数'
        ws1.cell(3, 1).value = '专业班级'
        ws1.cell(3, 2).value = '姓名'
        ws1.cell(3, 3).value = '人员类型'
        ws1.cell(3, 4).value = '学号'
        ws1.cell(3, 5).value = '性别'
        ws1.cell(3, 6).value = '年级'
        ws1.cell(3, 7).value = '手机号码'
        ws1.cell(3, 8).value = '累计次数'
        ws1.cell(3, 9).value = '未打卡日期'
        pointer = 4

        for e in printList:
            ws1.cell(pointer, 1).value = e['班级']
            ws1.cell(pointer, 2).value = e['姓名']
            ws1.cell(pointer, 3).value = e['人员类型']
            ws1.cell(pointer, 4).value = e['工号/学号']
            ws1.cell(pointer, 5).value = e['性别']
            ws1.cell(pointer, 6).value = e['年级']
            ws1.cell(pointer, 7).value = e['手机号']
            ws1.cell(pointer, 8).value = e['次数']
            ws1.cell(pointer, 9).value = ','.join(e['时间'])

            pointer += 1
        wb2.save(f"./yiban/output/通报批评汇总表（{timeRange[0]}-{timeRange[-1]}）.xlsx")
        pass


